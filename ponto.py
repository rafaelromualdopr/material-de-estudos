import openpyxl
from datetime import datetime, timedelta
import random

def generate_time(start_h, start_m, end_h, end_m):
    start = start_h * 60 + start_m
    end = end_h * 60 + end_m
    minutes = random.randint(start, end)
    return datetime(2025, 1, 1, minutes // 60, minutes % 60).strftime("%H:%M")

def calculate_exit(entrada, saida_almoco, entrada_retorno, jornada=6):
    periodo_manha = (datetime.strptime(saida_almoco, "%H:%M") - 
                    datetime.strptime(entrada, "%H:%M"))
    tempo_restante = timedelta(hours=jornada) - periodo_manha
    saida_final = (datetime.strptime(entrada_retorno, "%H:%M") + 
                  tempo_restante)
    return saida_final.strftime("%H:%M")

# Carregar a planilha
wb = openpyxl.load_workbook('W3 - Controle de ponto RAFAEL 02-25.xlsx')
sheet = wb['Controle de Ponto']

# Percorrer as linhas da planilha
for row in range(13, 41):  # Linhas com registros de ponto
    semana = sheet.cell(row=row, column=2).value
    
    # Pular finais de semana e dias não úteis
    if not semana or isinstance(semana, str) or semana in ['Sáb', 'Dom']:
        continue
        
    # Gerar horários
    entrada = generate_time(10, 0, 11, 30)
    saida_almoco = generate_time(12, 20, 13, 20)
    entrada_retorno = (datetime.strptime(saida_almoco, "%H:%M") + 
                      timedelta(hours=1)).strftime("%H:%M")
    saida_final = calculate_exit(entrada, saida_almoco, entrada_retorno)
    
    # Preencher as células
    sheet.cell(row=row, column=3, value=entrada)
    sheet.cell(row=row, column=4, value=saida_almoco)
    sheet.cell(row=row, column=5, value=entrada_retorno)
    sheet.cell(row=row, column=6, value=saida_final)

# Salvar a planilha modificada
wb.save('planilha_ponto_mensal.xlsx')
print("Planilha gerada com sucesso!")